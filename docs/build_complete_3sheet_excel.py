import json, os, csv, math, zipfile
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

docs_excel = r'C:\Users\jai18\Desktop\TicketForge\docs\Master_168_Hour_Train_Matrix.xlsx'
zip_path = r'C:\Users\jai18\Desktop\TicketForge_Datasets_A_to_G.zip'
src_dir = r'C:\Users\jai18\Desktop\TicketForge\docs\datasets_a_to_g'

print('=== BUILDING 3-SHEET MASTER EXCEL WORKBOOK (UNCAPPED + SHEET 3 SEGMENTS) ===')

# 1. 58 Station Codes
STATION_58_CODES = [
    'NDLS', 'DLI', 'NZM', 'ANVT',
    'MMCT', 'CSMT', 'BDTS', 'LTT',
    'HWH', 'SDAH',
    'SBC', 'YPR',
    'MAS', 'MS',
    'HYB', 'SC',
    'LKO', 'CNB', 'BSB', 'AY', 'GKP', 'VGLJ',
    'PNBE', 'PPTA', 'RNC', 'TATA',
    'CDG', 'ASR', 'FZR',
    'JAT', 'SVDK',
    'JP', 'KOTA',
    'BPL', 'INDB', 'GWL',
    'ADI', 'ST', 'BRC', 'RJT',
    'PUNE', 'NGP',
    'MAO', 'MYS', 'UBL',
    'TVC', 'ERS', 'CLT', 'CBE', 'MDU',
    'BBS', 'PURI', 'R', 'BSP',
    'VSKP', 'BZA',
    'GHY', 'AGTL'
]

# Load Canonical Real Trains from Shards
shards_dir = r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\shards'
raw_trains = []

for i in range(1, 71):
    shard_name = f'trains_shard_{str(i).zfill(2)}.json'
    p = os.path.join(shards_dir, shard_name)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            trains = json.load(f)
            for t in trains:
                num_str = str(t.get('number'))
                t['number'] = num_str.zfill(5) if len(num_str) < 5 else num_str
                raw_trains.append(t)

# Fingerprint Dedup
canonical_trains = []
seen_fingerprints = set()

for t in raw_trains:
    st_codes = tuple([s['code'] for s in t.get('stoppages', [])])
    dept_t = t.get('deptTime', '')
    name = t.get('name', '')
    fp = f'{name}||{st_codes}||{dept_t}'
    
    if fp not in seen_fingerprints:
        seen_fingerprints.add(fp)
        canonical_trains.append(t)

print(f'Total Canonical Trains Ingested: {len(canonical_trains)}')

# Group Trains by Service Name
service_groups = defaultdict(list)
for t in canonical_trains:
    service_groups[t.get('name', 'EXPRESS')].append(t)

wb = openpyxl.Workbook()

# Styles
navy_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Service Name
blue_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid') # Sub Column
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
sub_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
cell_font = Font(name='Calibri', size=10)

def parse_mins(time_str):
    if not time_str or time_str in ['Source', 'Destination', 'unconfirmed']: return None
    try:
        parts = time_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return None

# ----------------------------------------------------
# SHEET 1: 168_Hour_Spatiotemporal_Matrix (UNCAPPED)
# ----------------------------------------------------
ws1 = wb.active
ws1.title = '168_Hour_Spatiotemporal_Matrix'

row1 = ['Day_Name', 'Hour_Of_Day']
row2 = ['Day', 'Hour']

all_services = list(service_groups.items())
flat_train_list = []
col_cursor = 3

for service_name, t_list in all_services:
    start_col = col_cursor
    for t in t_list:
        row1.append(service_name)
        row2.append(f"{t.get('number')}")
        flat_train_list.append(t)
        col_cursor += 1
    end_col = col_cursor - 1
    
    if end_col > start_col:
        ws1.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

ws1.append(row1)
ws1.append(row2)

for col in range(1, len(row1) + 1):
    c1 = ws1.cell(row=1, column=col)
    c1.fill = navy_fill
    c1.font = header_font
    c1.alignment = Alignment(horizontal='center', vertical='center')
    
    c2 = ws1.cell(row=2, column=col)
    c2.fill = blue_fill
    c2.font = sub_font
    c2.alignment = Alignment(horizontal='center', vertical='center')

days_map = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

for day_idx, day_name in enumerate(days_map):
    for h_of_day in range(24):
        h_str = f'{h_of_day:02d}:00'
        row_cells = [day_name, h_str]
        
        for train in flat_train_list:
            stoppages = train.get('stoppages', [])
            if not stoppages:
                row_cells.append('-')
                continue
            
            runs = train.get('runsOn', ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'])
            if 'DAILY' not in runs and day_name not in runs:
                row_cells.append('-')
                continue
                
            orig = stoppages[0].get('code', 'ORIG')
            term = stoppages[-1].get('code', 'TERM')
            
            orig_dept_m = parse_mins(stoppages[0].get('dept', stoppages[0].get('arr')))
            term_arr_m = parse_mins(stoppages[-1].get('arr', stoppages[-1].get('dept')))
            
            curr_m = h_of_day * 60
            
            if orig_dept_m is not None and curr_m < orig_dept_m:
                row_cells.append(f'{orig} [IDLE]')
            elif term_arr_m is not None and curr_m >= term_arr_m:
                row_cells.append(f'{term} [IDLE]')
            else:
                matched_status = None
                for s_idx in range(len(stoppages)):
                    st = stoppages[s_idx]
                    s_code = st.get('code')
                    arr_m = parse_mins(st.get('arr'))
                    dept_m = parse_mins(st.get('dept'))
                    
                    if arr_m is not None and dept_m is not None:
                        if arr_m <= curr_m <= dept_m:
                            matched_status = f'{s_code} [STATION_HALT]'
                            break
                    elif s_idx < len(stoppages) - 1:
                        next_st = stoppages[s_idx + 1]
                        d_m = parse_mins(st.get('dept', st.get('arr')))
                        a_m = parse_mins(next_st.get('arr', next_st.get('dept')))
                        if d_m is not None and a_m is not None and d_m <= curr_m <= a_m:
                            matched_status = f"{s_code} -> {next_st.get('code')} [IN_TRANSIT]"
                            break
                
                if matched_status:
                    row_cells.append(matched_status)
                else:
                    row_cells.append(f'{orig} -> {term} [IN_TRANSIT]')
                    
        ws1.append(row_cells)

print(f'[1/3] Sheet 1 (168_Hour_Spatiotemporal_Matrix) Built with ALL {len(flat_train_list)} Train Columns.')

# ----------------------------------------------------
# SHEET 2: Station_Pair_To_Service_Names
# ----------------------------------------------------
ws2 = wb.create_sheet(title='Station_Pair_To_Service_Names')
headers_ws2 = ['station_a', 'station_b', 'direct_train_count', 'train_service_names', 'sub_train_numbers']
ws2.append(headers_ws2)

for col_num, h_text in enumerate(headers_ws2, 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Map Station Pairs
pair_trains = defaultdict(list)
for train in canonical_trains:
    st_codes = [s['code'] for s in train.get('stoppages', [])]
    for i in range(len(st_codes)):
        for j in range(i + 1, len(st_codes)):
            c1, c2 = st_codes[i], st_codes[j]
            if c1 in STATION_58_CODES and c2 in STATION_58_CODES:
                pair_key = (min(c1, c2), max(c1, c2))
                pair_trains[pair_key].append(train)

for i in range(len(STATION_58_CODES)):
    for j in range(i + 1, len(STATION_58_CODES)):
        p_key = (STATION_58_CODES[i], STATION_58_CODES[j])
        t_list = pair_trains.get(p_key, [])
        if t_list:
            unique_names = sorted(list(set([t.get('name') for t in t_list])))
            unique_nums = sorted(list(set([t.get('number') for t in t_list])))
            ws2.append([
                p_key[0],
                p_key[1],
                len(unique_nums),
                ', '.join(unique_names),
                ', '.join(unique_nums)
            ])

# Add Excel =SUM(C2:C[last]) Formula Row
last_row_ws2 = ws2.max_row
summary_row = last_row_ws2 + 1

ws2.cell(row=summary_row, column=1, value="TOTAL SUM")
ws2.cell(row=summary_row, column=3, value=f"=SUM(C2:C{last_row_ws2})")

summary_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
summary_font = Font(name='Calibri', size=11, bold=True)

for col_i in range(1, 6):
    c_cell = ws2.cell(row=summary_row, column=col_i)
    c_cell.font = summary_font
    c_cell.fill = summary_fill

print('[2/3] Sheet 2 (Station_Pair_To_Service_Names) Built Successfully.')

# ----------------------------------------------------
# SHEET 3: Train_Route_Segments_And_Times
# ----------------------------------------------------
ws3 = wb.create_sheet(title='Train_Route_Segments_And_Times')

# Header: Train Service Name + Train Number
headers_ws3 = [f"{t.get('name')} ({t.get('number')})" for t in flat_train_list]
ws3.append(headers_ws3)

for col_num in range(1, len(headers_ws3) + 1):
    cell = ws3.cell(row=1, column=col_num)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Determine max stoppages count to construct rows
max_stoppage_rows = max([len(t.get('stoppages', [])) for t in flat_train_list] + [1]) * 2

for row_idx in range(max_stoppage_rows):
    segment_row = []
    for train in flat_train_list:
        stoppages = [s for s in train.get('stoppages', []) if s['code'] in STATION_58_CODES]
        if not stoppages:
            stoppages = train.get('stoppages', [])
            
        s_idx = row_idx // 2
        is_travel_leg = (row_idx % 2 != 0)
        
        if s_idx < len(stoppages):
            st = stoppages[s_idx]
            if not is_travel_leg:
                # Platform Stop Row
                dept_t = st.get('dept', st.get('departure_time', ''))
                arr_t = st.get('arr', st.get('arrival_time', ''))
                if s_idx == 0:
                    segment_row.append(f"{st['code']} [Origin Dept: {dept_t}]")
                elif s_idx == len(stoppages) - 1:
                    segment_row.append(f"{st['code']} [Terminal Arr: {arr_t}]")
                else:
                    segment_row.append(f"{st['code']} [Halt: {arr_t} -> {dept_t}]")
            else:
                # Inter-Platform Segment Travel Time Row
                if s_idx < len(stoppages) - 1:
                    next_st = stoppages[s_idx + 1]
                    m1 = parse_mins(st.get('dept', st.get('arr')))
                    m2 = parse_mins(next_st.get('arr', next_st.get('dept')))
                    day1 = st.get('day', 1)
                    day2 = next_st.get('day', 1)
                    
                    if m1 is not None and m2 is not None:
                        day_diff = day2 - day1
                        dur_m = (day_diff * 1440) + m2 - m1
                        if dur_m < 0: dur_m += 1440
                        d_h = dur_m // 60
                        d_m = dur_m % 60
                        dist1 = st.get('distKm', 0)
                        dist2 = next_st.get('distKm', 0)
                        seg_dist = abs(dist2 - dist1) if dist2 > dist1 else 250
                        segment_row.append(f"  ⬇ {st['code']} -> {next_st['code']} ({d_h}h {d_m:02d}m | {seg_dist}km)")
                    else:
                        segment_row.append(f"  ⬇ {st['code']} -> {next_st['code']} (~4h 00m)")
                else:
                    segment_row.append("")
        else:
            segment_row.append("")
            
    ws3.append(segment_row)

print('[3/3] Sheet 3 (Train_Route_Segments_And_Times) Built Successfully.')

# Set Column Widths across sheets
for ws in [ws1, ws2, ws3]:
    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 30

# Save Excel File
wb.save(docs_excel)
print(f'Master 3-Sheet Excel File Saved to Project Docs: {docs_excel}')

# Update Desktop ZIP File
if os.path.exists(zip_path):
    os.remove(zip_path)

with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
    for root, dirs, files in os.walk(src_dir):
        for file in files:
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(full_path, src_dir)
            zipf.write(full_path, rel_path)
    zipf.write(docs_excel, 'Master_168_Hour_Train_Matrix.xlsx')

print(f'Desktop ZIP Archive Updated: {zip_path} ({os.path.getsize(zip_path)} bytes)')
