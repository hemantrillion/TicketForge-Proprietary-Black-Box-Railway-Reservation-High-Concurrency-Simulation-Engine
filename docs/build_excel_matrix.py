import json, os, csv, math
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

output_excel = r'C:\Users\jai18\Desktop\Master_168_Hour_Train_Matrix.xlsx'
docs_excel = r'C:\Users\jai18\Desktop\TicketForge\docs\Master_168_Hour_Train_Matrix.xlsx'

print('=== GENERATING 2-SHEET MASTER 168-HOUR EXCEL WORKBOOK ===')

# 1. 58 Station Codes
STATION_58_CODES = [
    'NDLS', 'DLI', 'NZM', 'ANVT',
    'MMCT', 'CSMT', 'BDTS', 'LTT',
    'HWH', 'SDAH',
    'SBC', 'YPR',
    'MAS', 'MS',
    'HYB', 'SC',
    'LKO', 'CNB', 'BSB', 'AY', 'GKP', 'VGLJ',
    'PNBE', 'PPTA', 'RNC', 'TATA',
    'CDG', 'ASR', 'FZR',
    'JAT', 'SVDK',
    'JP', 'KOTA',
    'BPL', 'INDB', 'GWL',
    'ADI', 'ST', 'BRC', 'RJT',
    'PUNE', 'NGP',
    'MAO', 'MYS', 'UBL',
    'TVC', 'ERS', 'CLT', 'CBE', 'MDU',
    'BBS', 'PURI', 'R', 'BSP',
    'VSKP', 'BZA',
    'GHY', 'AGTL'
]

# Load Canonical Real Trains from Shards
shards_dir = r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\shards'
raw_trains = []

for i in range(1, 71):
    shard_name = f'trains_shard_{str(i).zfill(2)}.json'
    p = os.path.join(shards_dir, shard_name)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            trains = json.load(f)
            for t in trains:
                num_str = str(t.get('number'))
                t['number'] = num_str.zfill(5) if len(num_str) < 5 else num_str
                raw_trains.append(t)

# Fingerprint Dedup
canonical_trains = []
seen_fingerprints = set()

for t in raw_trains:
    st_codes = tuple([s['code'] for s in t.get('stoppages', [])])
    dept_t = t.get('deptTime', '')
    name = t.get('name', '')
    fp = f'{name}||{st_codes}||{dept_t}'
    
    if fp not in seen_fingerprints:
        seen_fingerprints.add(fp)
        canonical_trains.append(t)

matrix_trains = canonical_trains[:80] # Clean representative trains column set

wb = openpyxl.Workbook()

# Setup Styles
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

# ----------------------------------------------------
# SHEET 1: 168_Hour_Spatiotemporal_Matrix
# ----------------------------------------------------
ws1 = wb.active
ws1.title = '168_Hour_Spatiotemporal_Matrix'

headers_ws1 = ['Day_Name', 'Hour_Of_Day'] + [f"{t.get('name')} ({t.get('number')})" for t in matrix_trains]
ws1.append(headers_ws1)

# Format Header
for col_num, h_text in enumerate(headers_ws1, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

days_map = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

def parse_mins(time_str):
    if not time_str or time_str in ['Source', 'Destination']: return None
    try:
        parts = time_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return None

# Populate 168 Hourly Rows (No Cell Empty)
for d_idx, day_name in enumerate(days_map):
    for h_of_day in range(24):
        time_str = f'{h_of_day:02d}:00'
        row_vals = [day_name, time_str]
        
        for t in matrix_trains:
            stoppages = t.get('stoppages', [])
            if not stoppages:
                row_vals.append('IDLE [No Stops]')
                continue
                
            orig_code = stoppages[0].get('code', 'ORIG')
            term_code = stoppages[-1].get('code', 'TERM')
            
            orig_dept_m = parse_mins(stoppages[0].get('dept', stoppages[0].get('arr')))
            term_arr_m = parse_mins(stoppages[-1].get('arr', stoppages[-1].get('dept')))
            
            h_mins = h_of_day * 60
            
            if orig_dept_m is not None and h_mins < orig_dept_m:
                row_vals.append(f'{orig_code} [Idle at Origin]')
            elif term_arr_m is not None and h_mins >= term_arr_m:
                row_vals.append(f'{term_code} [Idle at Terminal]')
            else:
                matched_status = None
                for idx_s in range(len(stoppages) - 1):
                    s1, s2 = stoppages[idx_s], stoppages[idx_s + 1]
                    m1 = parse_mins(s1.get('dept', s1.get('arr')))
                    m2 = parse_mins(s2.get('arr', s2.get('dept')))
                    
                    if m1 is not None and m2 is not None:
                        if m1 <= h_mins <= m2:
                            matched_status = f"{s1.get('code')} -> {s2.get('code')} [In-Transit]"
                            break
                
                if matched_status:
                    row_vals.append(matched_status)
                else:
                    row_vals.append(f"{orig_code} -> {term_code} [En-Route]")
                    
        ws1.append(row_vals)

print('[1/2] Sheet 1 (168_Hour_Spatiotemporal_Matrix) Built Successfully.')

# ----------------------------------------------------
# SHEET 2: Station_Pair_To_Service_Names
# ----------------------------------------------------
ws2 = wb.create_sheet(title='Station_Pair_To_Service_Names')

headers_ws2 = ['Station_A', 'Station_B', 'Direct_Train_Count', 'Train_Service_Names', 'Sub_Train_Numbers']
ws2.append(headers_ws2)

for col_num, h_text in enumerate(headers_ws2, 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Map Station Pairs
pair_trains = defaultdict(list)
for train in canonical_trains:
    st_codes = [s['code'] for s in train.get('stoppages', [])]
    for i in range(len(st_codes)):
        for j in range(i + 1, len(st_codes)):
            c1, c2 = st_codes[i], st_codes[j]
            if c1 in STATION_58_CODES and c2 in STATION_58_CODES:
                pair_key = (min(c1, c2), max(c1, c2))
                pair_trains[pair_key].append(train)

# Direct Pairs List
for i in range(len(STATION_58_CODES)):
    for j in range(i + 1, len(STATION_58_CODES)):
        p_key = (STATION_58_CODES[i], STATION_58_CODES[j])
        t_list = pair_trains.get(p_key, [])
        if t_list:
            unique_names = sorted(list(set([t.get('name') for t in t_list])))
            unique_nums = sorted(list(set([t.get('number') for t in t_list])))
            ws2.append([
                p_key[0],
                p_key[1],
                len(unique_nums),
                ', '.join(unique_names),
                ', '.join(unique_nums)
            ])

print('[2/2] Sheet 2 (Station_Pair_To_Service_Names) Built Successfully.')

# Save Excel Files
wb.save(output_excel)
wb.save(docs_excel)

print(f'Master Excel File Saved to Desktop: {output_excel}')
print(f'File Size: {os.path.getsize(output_excel)} bytes')
