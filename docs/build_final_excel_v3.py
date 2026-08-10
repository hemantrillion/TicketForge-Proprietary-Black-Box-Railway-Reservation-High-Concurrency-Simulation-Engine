import json, os, math, zipfile
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

docs_excel = r'C:\Users\jai18\Desktop\TicketForge\docs\Master_168_Hour_Train_Matrix.xlsx'
zip_path = r'C:\Users\jai18\Desktop\TicketForge_Datasets_A_to_G.zip'
src_dir = r'C:\Users\jai18\Desktop\TicketForge\docs\datasets_a_to_g'

print('=== REBUILDING 3-SHEET EXCEL WITH FULL STATION NAMES + CORRECT STATUS LOGIC ===')

# -------------------------------------------
# STATION NAME LOOKUP MAP (Code -> Full Name)
# -------------------------------------------
with open(r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\stations_57.json', 'r', encoding='utf-8') as f:
    st_data = json.load(f)

STATION_NAME = {s['code']: s['name'].split(' - ', 1)[1] if ' - ' in s['name'] else s['name'] for s in st_data}

# Add PPTA manually since it may not be in stations_57.json
STATION_NAME.setdefault('PPTA', 'Pataliputra Jn')

STATION_58_CODES = list(STATION_NAME.keys())

def sname(code):
    return STATION_NAME.get(code, code)

# Load Canonical Real Trains from Shards
shards_dir = r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\shards'
raw_trains = []

for i in range(1, 71):
    shard_name = f'trains_shard_{str(i).zfill(2)}.json'
    p = os.path.join(shards_dir, shard_name)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            trains = json.load(f)
            for t in trains:
                num_str = str(t.get('number'))
                t['number'] = num_str.zfill(5) if len(num_str) < 5 else num_str
                raw_trains.append(t)

# Fingerprint Dedup
canonical_trains = []
seen_fps = set()

for t in raw_trains:
    st_codes = tuple([s['code'] for s in t.get('stoppages', [])])
    fp = f"{t.get('name')}||{st_codes}||{t.get('deptTime','')}"
    if fp not in seen_fps:
        seen_fps.add(fp)
        canonical_trains.append(t)

print(f'Total Canonical Trains: {len(canonical_trains)}')

service_groups = defaultdict(list)
for t in canonical_trains:
    service_groups[t.get('name', 'EXPRESS')].append(t)

# Parse time to minutes
def parse_mins(time_str):
    if not time_str or time_str in ['Source', 'Destination', 'unconfirmed']: return None
    try:
        parts = time_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return None

# Build full timeline for a train: list of (global_min_start, global_min_end, status_label)
def build_train_timeline(train):
    """
    Returns a list of intervals:
    (start_min_from_midnight, end_min_from_midnight, label)
    Uses the day field in stoppages for multi-day trains.
    """
    stoppages = train.get('stoppages', [])
    timeline = []

    if not stoppages:
        return timeline

    for idx, st in enumerate(stoppages):
        code = st.get('code', '?')
        day_offset = (st.get('day', 1) - 1) * 1440  # convert day 1->0 mins, day 2->1440 mins, etc.
        arr_m = parse_mins(st.get('arr'))
        dept_m = parse_mins(st.get('dept'))

        if idx == 0:
            # Origin: no arrival, only depart
            if dept_m is not None:
                global_dept = day_offset + dept_m
                # Idle at origin before departure (from 0 to global_dept)
                if global_dept > 0:
                    timeline.append((0, global_dept, f'IDLE at {sname(code)} (Origin)'))
                timeline.append((global_dept, global_dept, f'DEPARTED from {sname(code)} (Origin)'))
        elif idx == len(stoppages) - 1:
            # Terminal: no departure
            if arr_m is not None:
                global_arr = day_offset + arr_m
                timeline.append((global_arr, global_arr, f'ARRIVED at {sname(code)} (Terminal)'))
                # Idle at terminal after arrival
                timeline.append((global_arr, 10080, f'IDLE at {sname(code)} (Terminal)'))
        else:
            # Intermediate: arrive then depart
            if arr_m is not None:
                global_arr = day_offset + arr_m
                timeline.append((global_arr, global_arr, f'ARRIVED at {sname(code)}'))
            if arr_m is not None and dept_m is not None:
                global_arr = day_offset + arr_m
                global_dept = day_offset + dept_m
                if global_dept > global_arr:
                    timeline.append((global_arr, global_dept, f'HALTED at {sname(code)} ({st.get("arr","")} - {st.get("dept","")})'))
                timeline.append((global_dept, global_dept, f'DEPARTED from {sname(code)}'))

        # In-transit to next stop
        if idx < len(stoppages) - 1:
            next_st = stoppages[idx + 1]
            next_day_offset = (next_st.get('day', 1) - 1) * 1440
            my_dept_m = parse_mins(st.get('dept', st.get('arr')))
            next_arr_m = parse_mins(next_st.get('arr', next_st.get('dept')))

            if my_dept_m is not None and next_arr_m is not None:
                g_dept = day_offset + my_dept_m
                g_next_arr = next_day_offset + next_arr_m
                if g_next_arr > g_dept:
                    timeline.append((g_dept, g_next_arr,
                        f'TRAVELLING {sname(code)} → {sname(next_st.get("code","?"))}'))

    # Sort by start time
    timeline.sort(key=lambda x: x[0])
    return timeline

def get_status_at_hour(timeline, day_idx, hour):
    """
    day_idx: 0=MON, 1=TUE, ... 6=SUN
    hour: 0-23
    Returns the status label at that exact hour.
    """
    global_min = day_idx * 1440 + hour * 60

    best = None
    for (start, end, label) in timeline:
        if start <= global_min <= end:
            best = label
            # Prefer more specific statuses
            if 'TRAVELLING' in label or 'HALTED' in label:
                return label
    
    if best:
        return best

    # If before first event
    if timeline and global_min < timeline[0][0]:
        return f'IDLE at {timeline[0][2].split(" ")[-1].strip("()")} (Pre-Service)'
    
    # If after last event
    if timeline and global_min > timeline[-1][1]:
        return f'IDLE (Off-Service)'

    return '-'


wb = openpyxl.Workbook()

# Styles
navy_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
blue_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
header_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
sub_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
day_font = Font(name='Calibri', size=12, bold=True)

# ============================================================
# SHEET 1: 168_Hour_Spatiotemporal_Matrix (UNCAPPED + FIXED)
# ============================================================
ws1 = wb.active
ws1.title = '168_Hour_Spatiotemporal_Matrix'

all_services = list(service_groups.items())
flat_train_list = []
row1 = ['Day_Name', 'Hour_Of_Day']
row2 = ['Day', 'Hour']
col_cursor = 3

for service_name, t_list in all_services:
    start_col = col_cursor
    for t in t_list:
        row1.append(service_name)
        row2.append(t.get('number'))
        flat_train_list.append(t)
        col_cursor += 1
    end_col = col_cursor - 1
    if end_col > start_col:
        ws1.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

ws1.append(row1)
ws1.append(row2)

# Row Heights
ws1.row_dimensions[1].height = 45
ws1.row_dimensions[2].height = 32

# Format Header Rows
for col in range(1, len(row1) + 1):
    c1 = ws1.cell(row=1, column=col)
    c1.fill = navy_fill
    c1.font = header_font
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    c2 = ws1.cell(row=2, column=col)
    c2.fill = blue_fill
    c2.font = sub_font
    c2.alignment = Alignment(horizontal='center', vertical='center')

# Pre-build timelines for all trains
print('Pre-building train timelines...')
timelines = {t['number']: build_train_timeline(t) for t in flat_train_list}

days_map = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

for day_idx, day_name in enumerate(days_map):
    for h_of_day in range(24):
        h_str = f'{h_of_day:02d}:00'
        row_cells = [day_name, h_str]

        for train in flat_train_list:
            runs = train.get('runsOn', ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'])
            if 'DAILY' not in runs and day_name not in runs:
                row_cells.append('-')
                continue

            tl = timelines.get(train['number'], [])
            status = get_status_at_hour(tl, day_idx, h_of_day)
            row_cells.append(status)

        ws1.append(row_cells)
        ws1.row_dimensions[ws1.max_row].height = 20

print(f'[1/3] Sheet 1 Built: ALL {len(flat_train_list)} Train Columns, Full Status Descriptions.')

# ============================================================
# SHEET 2: Station_Pair_To_Service_Names
# ============================================================
ws2 = wb.create_sheet(title='Station_Pair_To_Service_Names')
headers_ws2 = ['station_a', 'station_b', 'direct_train_count', 'train_service_names', 'sub_train_numbers']
ws2.append(headers_ws2)

ws2.row_dimensions[1].height = 32
for col_num in range(1, 6):
    cell = ws2.cell(row=1, column=col_num)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

pair_trains = defaultdict(list)
for train in canonical_trains:
    st_codes = [s['code'] for s in train.get('stoppages', [])]
    for i in range(len(st_codes)):
        for j in range(i + 1, len(st_codes)):
            c1, c2 = st_codes[i], st_codes[j]
            if c1 in STATION_58_CODES and c2 in STATION_58_CODES:
                pair_key = (min(c1, c2), max(c1, c2))
                pair_trains[pair_key].append(train)

for i in range(len(STATION_58_CODES)):
    for j in range(i + 1, len(STATION_58_CODES)):
        p_key = (STATION_58_CODES[i], STATION_58_CODES[j])
        t_list = pair_trains.get(p_key, [])
        if t_list:
            unique_names = sorted(list(set([t.get('name') for t in t_list])))
            unique_nums = sorted(list(set([t.get('number') for t in t_list])))
            ws2.append([p_key[0], p_key[1], len(unique_nums), ', '.join(unique_names), ', '.join(unique_nums)])

last_row_ws2 = ws2.max_row
summary_row = last_row_ws2 + 1
ws2.cell(row=summary_row, column=1, value='TOTAL SUM')
ws2.cell(row=summary_row, column=3, value=f'=SUM(C2:C{last_row_ws2})')
summary_font = Font(name='Calibri', size=11, bold=True)
summary_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
for col_i in range(1, 6):
    c = ws2.cell(row=summary_row, column=col_i)
    c.font = summary_font
    c.fill = summary_fill

print('[2/3] Sheet 2 Built Successfully.')

# ============================================================
# SHEET 3: Train_Route_Segments_And_Times (Full Names)
# ============================================================
ws3 = wb.create_sheet(title='Train_Route_Segments_And_Times')

headers_ws3 = [f"{t.get('name')} ({t.get('number')})" for t in flat_train_list]
ws3.append(headers_ws3)

ws3.row_dimensions[1].height = 45
for col_num in range(1, len(headers_ws3) + 1):
    cell = ws3.cell(row=1, column=col_num)
    cell.fill = navy_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

max_stoppage_rows = max([len(t.get('stoppages', [])) for t in flat_train_list] + [1]) * 2

for row_idx in range(max_stoppage_rows):
    segment_row = []
    for train in flat_train_list:
        stoppages = [s for s in train.get('stoppages', []) if s['code'] in STATION_58_CODES]
        if not stoppages:
            stoppages = train.get('stoppages', [])

        s_idx = row_idx // 2
        is_travel_leg = (row_idx % 2 != 0)

        if s_idx < len(stoppages):
            st = stoppages[s_idx]
            code = st.get('code', '?')
            full_name = sname(code)

            if not is_travel_leg:
                dept_t = st.get('dept', '')
                arr_t = st.get('arr', '')
                if s_idx == 0:
                    segment_row.append(f'{full_name} [Origin Dept: {dept_t}]')
                elif s_idx == len(stoppages) - 1:
                    segment_row.append(f'{full_name} [Terminal Arr: {arr_t}]')
                else:
                    segment_row.append(f'{full_name} [Halt: {arr_t} → {dept_t}]')
            else:
                if s_idx < len(stoppages) - 1:
                    next_st = stoppages[s_idx + 1]
                    m1 = parse_mins(st.get('dept', st.get('arr')))
                    m2 = parse_mins(next_st.get('arr', next_st.get('dept')))
                    day1 = st.get('day', 1)
                    day2 = next_st.get('day', 1)

                    if m1 is not None and m2 is not None:
                        day_diff = day2 - day1
                        dur_m = (day_diff * 1440) + m2 - m1
                        if dur_m < 0: dur_m += 1440
                        d_h = dur_m // 60
                        d_m = dur_m % 60
                        dist1 = st.get('distKm', 0)
                        dist2 = next_st.get('distKm', 0)
                        seg_dist = abs(dist2 - dist1) if dist2 > dist1 else '?'
                        segment_row.append(f'  ↓ → {sname(next_st.get("code","?"))} ({d_h}h {d_m:02d}m | {seg_dist}km)')
                    else:
                        segment_row.append(f'  ↓ → {sname(next_st.get("code","?"))}')
                else:
                    segment_row.append('')
        else:
            segment_row.append('')

    ws3.append(segment_row)

print('[3/3] Sheet 3 Built with Full Station Names & Exact Segment Times.')

# ============================================================
# SET COLUMN WIDTHS (Wide enough for full text)
# ============================================================
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 10
for col_idx in range(3, ws1.max_column + 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 38

for col_idx in range(1, ws2.max_column + 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35

ws3.column_dimensions['A'].width = 12
for col_idx in range(1, ws3.max_column + 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 38

# Save
wb.save(docs_excel)
print(f'Excel Saved: {docs_excel} ({os.path.getsize(docs_excel)} bytes)')

# Update ZIP
if os.path.exists(zip_path):
    os.remove(zip_path)
with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
    for root, dirs, files in os.walk(src_dir):
        for file in files:
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(full_path, src_dir)
            zipf.write(full_path, rel_path)
    zipf.write(docs_excel, 'Master_168_Hour_Train_Matrix.xlsx')
print(f'Desktop ZIP Updated: {zip_path} ({os.path.getsize(zip_path)} bytes)')
