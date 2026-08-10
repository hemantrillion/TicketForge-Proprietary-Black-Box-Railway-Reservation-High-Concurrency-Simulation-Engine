import json, os, math, zipfile
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

docs_excel = r'C:\Users\jai18\Desktop\TicketForge\docs\Master_168_Hour_Train_Matrix.xlsx'
zip_path = r'C:\Users\jai18\Desktop\TicketForge_Datasets_A_to_G.zip'
src_dir = r'C:\Users\jai18\Desktop\TicketForge\docs\datasets_a_to_g'

print('=== BUILDING FINAL EXCEL v4: CONTINUOUS BIDIRECTIONAL + NO IDLE + BRACKETS ===')

# Station Name Lookup
with open(r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\stations_57.json', 'r', encoding='utf-8') as f:
    st_data = json.load(f)

STATION_NAME = {s['code']: s['name'].split(' - ', 1)[1] if ' - ' in s['name'] else s['name'] for s in st_data}
STATION_NAME.setdefault('PPTA', 'Pataliputra Jn')
STATION_58_CODES = list(STATION_NAME.keys())

def sname(code):
    return STATION_NAME.get(code, code)

# Load Canonical Real Trains
shards_dir = r'C:\Users\jai18\Desktop\TicketForge\platform\frontend\public\data\spatiotemporal\shards'
raw_trains = []

for i in range(1, 71):
    shard_name = f'trains_shard_{str(i).zfill(2)}.json'
    p = os.path.join(shards_dir, shard_name)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            trains = json.load(f)
            for t in trains:
                num_str = str(t.get('number'))
                t['number'] = num_str.zfill(5) if len(num_str) < 5 else num_str
                raw_trains.append(t)

canonical_trains = []
seen_fps = set()
for t in raw_trains:
    st_codes = tuple([s['code'] for s in t.get('stoppages', [])])
    fp = f"{t.get('name')}||{st_codes}||{t.get('deptTime','')}"
    if fp not in seen_fps:
        seen_fps.add(fp)
        canonical_trains.append(t)

print(f'Total Canonical Trains: {len(canonical_trains)}')

service_groups = defaultdict(list)
for t in canonical_trains:
    service_groups[t.get('name', 'EXPRESS')].append(t)

def parse_mins(time_str):
    if not time_str or time_str in ['Source', 'Destination', 'unconfirmed']: return None
    try:
        parts = time_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except:
        return None

def build_continuous_timeline(train):
    """
    Builds a 10080-minute (168-hour) continuous bidirectional timeline.
    Train goes A->B then reverses B->A then A->B etc. with no idle ever.
    All labels are in [brackets].
    """
    stoppages = train.get('stoppages', [])
    if not stoppages:
        return []

    # Build one-way segment list: list of (start_min, end_min, label)
    forward_segments = []
    
    # First calculate total journey time for one direction
    first_dept_m = None
    last_arr_m = None
    last_arr_day = 1

    for idx, st in enumerate(stoppages):
        day_offset = (st.get('day', 1) - 1) * 1440
        code = st.get('code', '?')

        if idx == 0:
            dept_m = parse_mins(st.get('dept', st.get('arr')))
            if dept_m is not None:
                first_dept_m = day_offset + dept_m
                # origin departure segment
                forward_segments.append((first_dept_m, first_dept_m + 1,
                    f'[DEPARTED from {sname(code)} – Origin]'))
        elif idx == len(stoppages) - 1:
            arr_m = parse_mins(st.get('arr', st.get('dept')))
            if arr_m is not None:
                last_arr_m = day_offset + arr_m
                last_arr_day = st.get('day', 1)
        else:
            arr_m = parse_mins(st.get('arr'))
            dept_m = parse_mins(st.get('dept'))
            if arr_m is not None:
                g_arr = day_offset + arr_m
                forward_segments.append((g_arr, g_arr + 1, f'[ARRIVED at {sname(code)}]'))
            if arr_m is not None and dept_m is not None:
                g_arr = day_offset + arr_m
                g_dept = day_offset + dept_m
                if g_dept > g_arr:
                    forward_segments.append((g_arr, g_dept,
                        f'[HALTED at {sname(code)} ({st.get("arr","")}–{st.get("dept","")})]'))
                forward_segments.append((g_dept, g_dept + 1, f'[DEPARTED from {sname(code)}]'))

        # In-transit to next
        if idx < len(stoppages) - 1:
            next_st = stoppages[idx + 1]
            next_day_offset = (next_st.get('day', 1) - 1) * 1440
            d_m = parse_mins(st.get('dept', st.get('arr')))
            a_m = parse_mins(next_st.get('arr', next_st.get('dept')))
            if d_m is not None and a_m is not None:
                g_d = day_offset + d_m
                g_a = next_day_offset + a_m
                if g_a > g_d:
                    forward_segments.append((g_d, g_a,
                        f'[TRAVELLING {sname(code)} → {sname(next_st.get("code","?"))}]'))

    if first_dept_m is None or last_arr_m is None:
        return []

    one_way_duration = last_arr_m - first_dept_m  # Minutes for full one-way trip
    if one_way_duration <= 0:
        one_way_duration = 720  # fallback 12h

    # Build reverse segments (B->A mirror)
    reverse_segments = []
    orig_code = stoppages[0].get('code', '?')
    term_code = stoppages[-1].get('code', '?')

    # Simple reverse: mirror the forward trip in reverse order
    rev_stoppages = list(reversed(stoppages))
    for idx, st in enumerate(rev_stoppages):
        code = st.get('code', '?')
        day_offset = (st.get('day', 1) - 1) * 1440
        rev_day_offset = (last_arr_day - 1 - st.get('day', 1) + last_arr_day) * 1440

        if idx < len(rev_stoppages) - 1:
            next_st = rev_stoppages[idx + 1]
            d_label = f'[TRAVELLING {sname(code)} → {sname(next_st.get("code","?"))} (Return)]'
            next_day_offset_r = (next_st.get('day', 1) - 1) * 1440
            d_m = parse_mins(st.get('arr', st.get('dept')))
            a_m = parse_mins(next_st.get('dept', next_st.get('arr')))
            if d_m is not None and a_m is not None:
                seg_dur = abs((day_offset + d_m) - (next_day_offset_r + a_m))
                if seg_dur <= 0:
                    seg_dur = 60
                reverse_segments.append((0, seg_dur, d_label))

    # Now stitch everything into a continuous 10080-min timeline
    timeline = []  # (absolute_minute_start, absolute_minute_end, label)

    # Build one full cycle = forward + reverse
    # Forward: from 0 -> one_way_duration
    for (rs, re, label) in forward_segments:
        timeline.append((rs, re, label))

    # Reverse trip starts at one_way_duration
    rev_cursor = last_arr_m
    for (_, dur, label) in reverse_segments:
        timeline.append((rev_cursor, rev_cursor + dur, label))
        rev_cursor += dur

    round_trip_duration = rev_cursor

    if round_trip_duration <= 0:
        round_trip_duration = one_way_duration * 2

    # Tile this round trip across the 10080 minutes
    full_timeline = []
    base_offset = 0

    while base_offset < 10080:
        for (s, e, label) in timeline:
            abs_s = s + base_offset
            abs_e = e + base_offset
            if abs_s >= 10080:
                break
            full_timeline.append((abs_s, min(abs_e, 10080), label))
        base_offset += round_trip_duration

    full_timeline.sort(key=lambda x: x[0])
    return full_timeline

def get_status_at_hour(timeline, day_idx, hour):
    """
    Checks the full 60-minute window [hour*60, (hour+1)*60] for any event.
    Priority: HALTED > ARRIVED > DEPARTED > TRAVELLING
    This ensures short stops (even 2-5 min halts) are captured in the correct hour slot.
    """
    global_min_start = day_idx * 1440 + hour * 60
    global_min_end   = global_min_start + 60

    best = None
    best_priority = -1

    for (start, end, label) in timeline:
        # Any overlap between segment [start,end] and window [global_min_start, global_min_end]
        if start <= global_min_end and end >= global_min_start:
            if 'HALTED' in label:
                priority = 5
            elif 'ARRIVED' in label:
                priority = 4
            elif 'DEPARTED' in label:
                priority = 3
            elif 'TRAVELLING' in label:
                priority = 2
            else:
                priority = 1

            if priority > best_priority:
                best = label
                best_priority = priority

    if best:
        return best

    # Nothing found in window — use next upcoming event
    for (start, end, label) in timeline:
        if start > global_min_end:
            return label

    return timeline[-1][2] if timeline else '[-]'



wb = openpyxl.Workbook()

# Styles - both rows Navy/Dark with White font
navy_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
header_font_r1 = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
header_font_r2 = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

# ============================================================
# SHEET 1: 168_Hour_Spatiotemporal_Matrix
# ============================================================
ws1 = wb.active
ws1.title = '168_Hour_Spatiotemporal_Matrix'

all_services = list(service_groups.items())
flat_train_list = []
row1 = ['Day_Name', 'Hour_Of_Day']
row2 = ['Day', 'Hour']
col_cursor = 3

for service_name, t_list in all_services:
    start_col = col_cursor
    for t in t_list:
        row1.append(service_name)
        row2.append(t.get('number'))
        flat_train_list.append(t)
        col_cursor += 1
    end_col = col_cursor - 1
    if end_col > start_col:
        ws1.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

ws1.append(row1)
ws1.append(row2)

# Row heights
ws1.row_dimensions[1].height = 45
ws1.row_dimensions[2].height = 32

# Both header rows: Navy fill + White font
for col in range(1, len(row1) + 1):
    c1 = ws1.cell(row=1, column=col)
    c1.fill = navy_fill
    c1.font = header_font_r1
    c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    c2 = ws1.cell(row=2, column=col)
    c2.fill = navy_fill  # Same navy, NOT blue
    c2.font = header_font_r2
    c2.alignment = Alignment(horizontal='center', vertical='center')

# Pre-build timelines
print('Pre-building continuous bidirectional train timelines...')
timelines = {}
for t in flat_train_list:
    timelines[t['number']] = build_continuous_timeline(t)

days_map = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

for day_idx, day_name in enumerate(days_map):
    for h_of_day in range(24):
        h_str = f'{h_of_day:02d}:00'
        row_cells = [day_name, h_str]

        for train in flat_train_list:
            runs = train.get('runsOn', ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN'])
            if 'DAILY' not in runs and day_name not in runs:
                row_cells.append('-')
                continue
            tl = timelines.get(train['number'], [])
            status = get_status_at_hour(tl, day_idx, h_of_day)
            row_cells.append(status)

        ws1.append(row_cells)

print(f'[1/3] Sheet 1: ALL {len(flat_train_list)} Columns, Continuous Bidirectional, No Idle.')

# ============================================================
# SHEET 2: Station_Pair_To_Service_Names
# ============================================================
ws2 = wb.create_sheet(title='Station_Pair_To_Service_Names')
headers_ws2 = ['station_a', 'station_b', 'direct_train_count', 'train_service_names', 'sub_train_numbers']
ws2.append(headers_ws2)
ws2.row_dimensions[1].height = 32
for col_num in range(1, 6):
    c = ws2.cell(row=1, column=col_num)
    c.fill = navy_fill
    c.font = header_font_r1
    c.alignment = Alignment(horizontal='center', vertical='center')

pair_trains = defaultdict(list)
for train in canonical_trains:
    st_codes = [s['code'] for s in train.get('stoppages', [])]
    for i in range(len(st_codes)):
        for j in range(i + 1, len(st_codes)):
            c1, c2 = st_codes[i], st_codes[j]
            if c1 in STATION_58_CODES and c2 in STATION_58_CODES:
                pair_key = (min(c1, c2), max(c1, c2))
                pair_trains[pair_key].append(train)

for i in range(len(STATION_58_CODES)):
    for j in range(i + 1, len(STATION_58_CODES)):
        p_key = (STATION_58_CODES[i], STATION_58_CODES[j])
        t_list = pair_trains.get(p_key, [])
        if t_list:
            unique_names = sorted(list(set([t.get('name') for t in t_list])))
            unique_nums = sorted(list(set([t.get('number') for t in t_list])))
            ws2.append([p_key[0], p_key[1], len(unique_nums), ', '.join(unique_names), ', '.join(unique_nums)])

last_row_ws2 = ws2.max_row
ws2.cell(row=last_row_ws2 + 1, column=1, value='TOTAL SUM')
ws2.cell(row=last_row_ws2 + 1, column=3, value=f'=SUM(C2:C{last_row_ws2})')
for col_i in range(1, 6):
    c = ws2.cell(row=last_row_ws2 + 1, column=col_i)
    c.font = Font(name='Calibri', size=11, bold=True)
    c.fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

print('[2/3] Sheet 2 Built.')

# ============================================================
# SHEET 3: Train_Route_Segments_And_Times
# ============================================================
ws3 = wb.create_sheet(title='Train_Route_Segments_And_Times')
headers_ws3 = [f"{t.get('name')} ({t.get('number')})" for t in flat_train_list]
ws3.append(headers_ws3)
ws3.row_dimensions[1].height = 45
for col_num in range(1, len(headers_ws3) + 1):
    c = ws3.cell(row=1, column=col_num)
    c.fill = navy_fill
    c.font = header_font_r1
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

max_stoppage_rows = max([len(t.get('stoppages', [])) for t in flat_train_list] + [1]) * 2

for row_idx in range(max_stoppage_rows):
    segment_row = []
    for train in flat_train_list:
        stoppages = [s for s in train.get('stoppages', []) if s['code'] in STATION_58_CODES]
        if not stoppages:
            stoppages = train.get('stoppages', [])

        s_idx = row_idx // 2
        is_travel_leg = (row_idx % 2 != 0)

        if s_idx < len(stoppages):
            st = stoppages[s_idx]
            code = st.get('code', '?')
            full_name = sname(code)

            if not is_travel_leg:
                dept_t = st.get('dept', '')
                arr_t = st.get('arr', '')
                if s_idx == 0:
                    segment_row.append(f'[{full_name} – Origin Dept: {dept_t}]')
                elif s_idx == len(stoppages) - 1:
                    segment_row.append(f'[{full_name} – Terminal Arr: {arr_t}]')
                else:
                    segment_row.append(f'[{full_name} – Halt: {arr_t} → {dept_t}]')
            else:
                if s_idx < len(stoppages) - 1:
                    next_st = stoppages[s_idx + 1]
                    m1 = parse_mins(st.get('dept', st.get('arr')))
                    m2 = parse_mins(next_st.get('arr', next_st.get('dept')))
                    day1 = st.get('day', 1)
                    day2 = next_st.get('day', 1)
                    if m1 is not None and m2 is not None:
                        dur_m = (day2 - day1) * 1440 + m2 - m1
                        if dur_m < 0: dur_m += 1440
                        d_h, d_m = dur_m // 60, dur_m % 60
                        dist1 = st.get('distKm', 0)
                        dist2 = next_st.get('distKm', 0)
                        seg_dist = abs(dist2 - dist1) if dist2 > dist1 else '?'
                        segment_row.append(f'  ↓ → [{sname(next_st.get("code","?"))} – {d_h}h {d_m:02d}m | {seg_dist}km]')
                    else:
                        segment_row.append(f'  ↓ → [{sname(next_st.get("code","?"))}]')
                else:
                    segment_row.append('')
        else:
            segment_row.append('')

    ws3.append(segment_row)

print('[3/3] Sheet 3 Built.')

# Column Widths
ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 10
for col_idx in range(3, ws1.max_column + 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40
for col_idx in range(1, ws2.max_column + 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35
for col_idx in range(1, ws3.max_column + 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 40

wb.save(docs_excel)
print(f'Excel Saved: {docs_excel} ({os.path.getsize(docs_excel)} bytes)')

if os.path.exists(zip_path):
    os.remove(zip_path)
with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
    for root, dirs, files in os.walk(src_dir):
        for file in files:
            full_path = os.path.join(root, file)
            zipf.write(full_path, os.path.relpath(full_path, src_dir))
    zipf.write(docs_excel, 'Master_168_Hour_Train_Matrix.xlsx')
print(f'ZIP Updated: {zip_path} ({os.path.getsize(zip_path)} bytes)')
